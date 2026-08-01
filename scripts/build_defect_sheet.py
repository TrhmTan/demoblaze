#!/usr/bin/env python3
"""
Add/rebuild the "Defect" sheet in inputdata/Demoblaze_QA_TestCases.xlsx.

This is curated data (defect descriptions, severity, root cause) - it doesn't
come from the Playwright JSON report the way Actual Result/Status do. What
this script DOES pull live from the workbook is the current Status of each
related TC ID (from the Login/Cart/API sheets, which sync_test_results_to_xlsx.py
keeps current), so the "Automation Evidence" column never goes stale relative
to whatever the last sync wrote.

Usage:
    python3 scripts/build_defect_sheet.py [--xlsx inputdata/Demoblaze_QA_TestCases.xlsx]

Safe to re-run: it fully replaces the "Defect" sheet each time rather than
appending, so editing DEFECTS below and re-running is the way to update it.
"""
import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency: run `pip install openpyxl` (or `pip install openpyxl --break-system-packages`) first.")
    sys.exit(1)

DEFECTS = [
    {
        "id": "DEF-SYS-001",
        "title": "Guest cart identity not established outside index.html -> shared-bucket data leak between anonymous sessions",
        "area": "Cart / Guest identity",
        "severity": "Critical",
        "related_tc": ["TC-CRT-046"],
        "root_cause": (
            "The guest-cart cookie (user=<uuid>) is only set by js/index.js on the "
            "homepage. Arriving via a direct/bookmarked product link (prod.html) or "
            "cart.html first never sets it, so /viewcart is called with an empty "
            "cookie and the API returns the bucket SHARED by every cookie-less guest "
            "on demoblaze.com - not an empty cart."
        ),
        "evidence": (
            "Verified growing over time from unrelated traffic: 168 items -> 589 "
            "items (next day) -> 720 items (day after). This is a live data leak "
            "between unrelated guest sessions, not stale test data."
        ),
        "reference": "reports/test-cases/TC-CRT-046-definition.md",
    },
    {
        "id": "DEF-001",
        "title": "Credit card field has no format/length/Luhn validation",
        "area": "Checkout",
        "severity": "Critical",
        "related_tc": ["TC-CRT-020", "TC-CRT-027", "TC-CRT-028", "TC-CRT-029", "TC-CRT-030"],
        "root_cause": (
            "purchaseOrder() submits the card field exactly as typed with no "
            "client- or server-side validation: non-numeric characters, spaces, "
            "wrong length (13-19 digits expected), and failing Luhn checksums are "
            "all accepted and the order still succeeds."
        ),
        "evidence": "",
        "reference": "reports/test-cases/financial-validation-rules-and-case-redefinition.md",
    },
    {
        "id": "DEF-002",
        "title": "Expiry month never validated",
        "area": "Checkout",
        "severity": "High",
        "related_tc": ["TC-CRT-031", "TC-CRT-032", "TC-CRT-033"],
        "root_cause": "Month value is accepted as-is: 0, 13, and non-numeric strings all pass through and the order still succeeds.",
        "evidence": "",
        "reference": "reports/test-cases/financial-validation-rules-and-case-redefinition.md",
    },
    {
        "id": "DEF-003",
        "title": "Expiry year never validated",
        "area": "Checkout",
        "severity": "High",
        "related_tc": ["TC-CRT-035", "TC-CRT-036", "TC-CRT-037", "TC-CRT-038"],
        "root_cause": (
            "Year value is accepted as-is: already-expired years, 2-digit years, "
            "far-future years, and non-numeric strings all pass through and the "
            "order still succeeds. TC-CRT-022 (older duplicate, past-year case) "
            "still asserts the old permissive behavior and has not been redefined "
            "to match this defect yet - see Notes on that row."
        ),
        "evidence": "",
        "reference": "reports/test-cases/financial-validation-rules-and-case-redefinition.md",
    },
    {
        "id": "DEF-004",
        "title": "Empty cart can still be \"placed\" as a 0 USD order",
        "area": "Checkout",
        "severity": "Medium",
        "related_tc": ["TC-CRT-023b"],
        "root_cause": (
            "Nothing disables the Place Order button or rejects submission when "
            "the cart has zero items; purchaseOrder() happily creates an order ID "
            "for 0 USD. TC-CRT-023 (older duplicate) still asserts the old "
            "permissive behavior."
        ),
        "evidence": "",
        "reference": "reports/test-cases/financial-validation-rules-and-case-redefinition.md",
    },
    {
        "id": "DEF-005",
        "title": "Purchase button stays clickable behind the SweetAlert success popup",
        "area": "Checkout",
        "severity": "Medium",
        "related_tc": ["TC-CRT-047"],
        "root_cause": (
            "The Place Order modal is not disabled/hidden when the SweetAlert "
            "success confirmation appears on top of it, so the Purchase button "
            "underneath remains visible and clickable - a second click while the "
            "confirmation is showing can plausibly fire a duplicate order."
        ),
        "evidence": "",
        "reference": "reports/test-cases/Implementation-checklist-for-manual-TC.md",
    },
    {
        "id": "DEF-006",
        "title": "Invoice date is off by one month on every order",
        "area": "Checkout",
        "severity": "Low",
        "related_tc": [],
        "root_cause": (
            "purchaseOrder() builds the invoice date with `date.getMonth()` "
            "without the customary +1 (getMonth() is 0-indexed in JavaScript), so "
            "every invoice prints the previous month. Present on every passing "
            "checkout test today - nothing currently asserts on the date, so it "
            "has gone unnoticed."
        ),
        "evidence": "Not automated yet - no TC ID/test case covers this.",
        "reference": "reports/test-cases/cart-failure-analysis.md",
    },
    {
        "id": "DEF-API-002",
        "title": "POST /login accepts invalid/malformed credentials without rejecting (no 401/400)",
        "area": "API / /login",
        "severity": "Critical",
        "related_tc": ["API-LOG-002", "API-LOG-004", "API-LOG-005", "API-LOG-006"],
        "root_cause": (
            "Wrong password, empty username, empty password, and SQL-injection-shaped input "
            "(' OR '1'='1) are all expected to return 401/400 with a structured error body, but "
            "the live API does not reject any of them the way a hardened auth endpoint should."
        ),
        "evidence": "",
        "reference": "docs/api-analysis/BUG_REPORT_API_FAILURES.md",
    },
    {
        "id": "DEF-API-003",
        "title": "Cart endpoints (/viewcart, /addtocart, /deleteitem) enforce no auth/ownership or input validation",
        "area": "API / cart endpoints",
        "severity": "Critical",
        "related_tc": ["API-CART-002", "API-CART-003", "API-ADD-003", "API-ADD-004", "API-ADD-005",
                       "API-DEL-002", "API-DEL-003"],
        "root_cause": (
            "Empty/missing cookie, invalid product ID, missing cookie on add, flag=false, and "
            "deleting a non-existent item are all expected to return 400/401/404, but the live "
            "API accepts or no-ops on all of them with HTTP 200 instead. Combined with "
            "DEF-SYS-001 (shared guest bucket), this confirms the cart endpoints do not treat "
            "the cookie as a real auth/ownership token."
        ),
        "evidence": "",
        "reference": "docs/api-analysis/BUG_REPORT_API_FAILURES.md",
    },
    {
        "id": "DEF-API-004",
        "title": "POST /view (product details) does not validate the product ID",
        "area": "API / /view",
        "severity": "High",
        "related_tc": ["API-PROD-003", "API-PROD-004", "API-PROD-005"],
        "root_cause": (
            "A non-existent product ID (99999), a missing idp_ field, and idp_=0 are all "
            "expected to return 400/404 with a structured error, but the live API returns 200 "
            "for all three instead of validating the ID."
        ),
        "evidence": "",
        "reference": "docs/api-analysis/BUG_REPORT_API_FAILURES.md",
    },
    {
        "id": "DEF-API-005",
        "title": "POST /purchaseorder accepts invalid financial data and missing required fields with no validation",
        "area": "API / /purchaseorder",
        "severity": "Critical",
        "related_tc": ["API-ORD-002", "API-ORD-003", "API-ORD-004", "API-ORD-005", "API-ORD-006",
                       "API-ORD-007", "API-ORD-008", "API-ORD-009", "API-ORD-010", "API-ORD-011",
                       "API-ORD-012", "API-ORD-013"],
        "root_cause": (
            "Empty cart, non-numeric/wrong-length card number, invalid/non-numeric/out-of-range "
            "month, expired/2-digit/far-future year, and missing required fields (country, "
            "cardnumber) are all expected to return 400 with a structured error, but the live "
            "checkout endpoint accepts all of them and returns 200 - this is the API-level "
            "confirmation of the same root cause as DEF-001/002/003/004 found at the UI layer."
        ),
        "evidence": "",
        "reference": "docs/api-analysis/BUG_REPORT_API_FAILURES.md",
    },
    {
        "id": "DEF-API-006",
        "title": "GET /config.json is publicly exposed (information disclosure)",
        "area": "API / config.json",
        "severity": "Low",
        "related_tc": ["API-CONF-001"],
        "root_cause": (
            "config.json is expected to 404 (it should not be a public endpoint), but it is "
            "reachable and returns 200. Minor info-disclosure finding rather than a functional "
            "bug - flagging separately from the validation-pattern defects above since the risk "
            "profile is different (exposure, not broken business logic)."
        ),
        "evidence": "",
        "reference": "docs/api-analysis/BUG_REPORT_API_FAILURES.md",
    },
    {
        "id": "DEF-TEST-002",
        "title": "4 API happy-path tests asserted a response shape/value the app never returns (test-script bugs, not app defects)",
        "area": "Test infrastructure (not an app defect)",
        "severity": "Low",
        "related_tc": ["API-CART-001", "API-ADD-002", "API-DEL-004", "API-PROD-001"],
        "root_cause": (
            "API-CART-001, API-ADD-002, and API-DEL-004 all asserted a made-up REST-ish response "
            "shape (`{status, data: {items}}`) for /viewcart, but the real, already-confirmed "
            "shape used successfully throughout CartPage.ts all along is `{Items: [...]}` - a "
            "top-level, capital-I array. Separately, API-PROD-001 asserted the product title as "
            "'Samsung Galaxy S6' (capital G) when the real catalog title, used successfully "
            "throughout cart.spec.ts/regression.spec.ts, is 'Samsung galaxy s6' (lowercase). Both "
            "were test-authoring assumptions written without checking the real API against known "
            "working code already in this repo. Fixed directly in tests/api.spec.ts."
        ),
        "evidence": "Fixed in this repo - re-run API suite to confirm these 4 now pass.",
        "reference": "tests/api.spec.ts",
    },
    {
        "id": "DEF-TEST-001",
        "title": "DEF-00x checkout test cases were hanging for ~96-100s instead of failing fast",
        "area": "Test infrastructure (not an app defect)",
        "severity": "Low",
        "related_tc": ["TC-CRT-020", "TC-CRT-027", "TC-CRT-028", "TC-CRT-029", "TC-CRT-030",
                       "TC-CRT-031", "TC-CRT-032", "TC-CRT-033", "TC-CRT-035", "TC-CRT-036",
                       "TC-CRT-037", "TC-CRT-038", "TC-CRT-023b"],
        "root_cause": (
            "clickPurchaseExpectingAlert() waits indefinitely for a native "
            "window.alert() to confirm the app rejected bad input. Since DEF-001/002/003 "
            "mean the app never shows that alert (it just submits the order instead), the "
            "wait never resolves and every one of these tests ran to the full 90-120s "
            "config timeout before failing - correct end result (Fail), wrong reason "
            "(timeout, not a clean assertion), and ~20+ minutes wasted per full run. "
            "Fixed by racing the alert against the SweetAlert success popup with a short "
            "bound, so these now fail in a few seconds with a clear message instead."
        ),
        "evidence": "Logged from reports/test-results/archive/2026-08-01-run-001/results.json (durations 95.7s-100.0s).",
        "reference": "reports/test-cases/cart-failure-analysis.md",
    },
]

HEADERS = [
    "Defect ID", "Title", "Area", "Severity", "Related TC IDs",
    "Root Cause", "Automation Evidence (live)", "Reference",
]

SEVERITY_FILL = {
    "Critical": "FFC7CE",
    "High": "FFE4B5",
    "Medium": "FFF2CC",
    "Low": "E2EFDA",
}


def read_tc_status(wb, tc_ids):
    """Look up each TC ID's current Status/Actual Result from the Login/Cart/API
    sheets so the Defect sheet's evidence column reflects the latest sync,
    not a hand-typed snapshot that will silently go stale."""
    status_by_tc = {}
    for sheet_name in ("Login", "Cart", "API"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row, headers = None, {}
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=r, column=c).value).strip() == "TC ID":
                    header_row = r
                    for cc in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=cc).value
                        if v:
                            headers[str(v).strip()] = cc
                    break
            if header_row:
                break
        if not header_row:
            continue
        col_tcid = headers.get("TC ID")
        col_status = headers.get("Status")
        for r in range(header_row + 1, ws.max_row + 1):
            tcid = ws.cell(row=r, column=col_tcid).value
            if tcid and str(tcid).strip() in tc_ids:
                status_by_tc[str(tcid).strip()] = ws.cell(row=r, column=col_status).value
    return status_by_tc


def build_evidence_text(defect, status_by_tc):
    if defect["evidence"]:
        base = defect["evidence"]
    else:
        base = ""
    if not defect["related_tc"]:
        return base or "No automated TC currently covers this."
    parts = [f"{tc}: {status_by_tc.get(tc, 'not found')}" for tc in defect["related_tc"]]
    joined = " | ".join(parts)
    return f"{base} ({joined})" if base else joined


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="inputdata/Demoblaze_QA_TestCases.xlsx")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"No xlsx file at {xlsx_path}.")
        sys.exit(1)

    backup_path = xlsx_path.with_name(
        f"{xlsx_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{xlsx_path.suffix}"
    )
    shutil.copy2(xlsx_path, backup_path)
    print(f"Backup written: {backup_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    all_related_tc = {tc for d in DEFECTS for tc in d["related_tc"]}
    status_by_tc = read_tc_status(wb, all_related_tc)

    if "Defect" in wb.sheetnames:
        del wb["Defect"]
    ws = wb.create_sheet("Defect", index=wb.sheetnames.index("API") + 1 if "API" in wb.sheetnames else len(wb.sheetnames))

    bold = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=1, column=1, value="Defect – Found while building the automation suite").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    r = 3
    for defect in DEFECTS:
        evidence = build_evidence_text(defect, status_by_tc)
        row_values = [
            defect["id"],
            defect["title"],
            defect["area"],
            defect["severity"],
            ", ".join(defect["related_tc"]) if defect["related_tc"] else "(none automated yet)",
            defect["root_cause"],
            evidence,
            defect["reference"],
        ]
        for c, v in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = wrap
        sev_cell = ws.cell(row=r, column=4)
        fill_color = SEVERITY_FILL.get(defect["severity"])
        if fill_color:
            sev_cell.fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=r, column=1).font = bold
        r += 1

    widths = [14, 46, 26, 10, 24, 55, 55, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")
    print(f"Defect sheet: {len(DEFECTS)} defects written.")


if __name__ == "__main__":
    main()
