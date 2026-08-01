#!/usr/bin/env python3
"""
Fill in the "Actual Result" and "Status" columns of
inputdata/Demoblaze_QA_TestCases.xlsx (sheets: Login, Cart, API) from a real
Playwright JSON test report.

Why this exists: playwright.config.ts only had html+list reporters (no
structured output, overwritten every run), so there was no machine-readable,
dated evidence to base "Actual Result"/"Status" on - every run's results
were gone as soon as the next run started. playwright.manual.config.ts now
writes reports/test-results/latest/results.json every time
`npm run test:manual-suite` runs, and scripts/archive-run.js preserves the
previous run under reports/test-results/archive/ first. This script reads
that JSON and writes it into the xlsx.

Mapping TC ID -> automated test: every test() title in login.spec.ts,
cart.spec.ts, api.spec.ts starts with its manual TC ID
(e.g. "TC-CRT-005: Delete a product from the cart"), so we match on that
prefix rather than test order/index.

Known duplicate TC IDs in cart.spec.ts (same ID, two tests - an original and
a later "[DEF-xxx]" retest written after a defect fix): TC-CRT-020, 027,
028, 029, 030, 031, 032. Per team decision, the "[DEF-xxx]" version is
authoritative for Actual Result/Status; the older duplicate is only
referenced in the Notes column so the history isn't silently dropped.

Usage:
    python3 scripts/sync_test_results_to_xlsx.py \
        [--results reports/test-results/latest/results.json] \
        [--xlsx inputdata/Demoblaze_QA_TestCases.xlsx]

This NEVER fabricates results: it only writes what is actually present in
the Playwright JSON report. TC IDs with no matching automated test, or not
present in this run's JSON, are left untouched (still "Not Run" / whatever
they already said).
"""
import argparse
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: run `pip install openpyxl` (or `pip install openpyxl --break-system-packages`) first.")
    sys.exit(1)

TC_ID_RE = re.compile(r"^(TC-[A-Z]+-\d+|API-[A-Z]+-\d+)")
ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


def strip_ansi(text):
    """Strip ANSI color codes from Playwright error messages - openpyxl
    refuses to write raw escape characters (IllegalCharacterError)."""
    return ANSI_RE.sub("", text) if text else text

FILE_TO_SHEET = {
    "login.spec.ts": "Login",
    "cart.spec.ts": "Cart",
    "api.spec.ts": "API",
}

STATUS_MAP = {
    # Playwright's per-test outcome status (test.status in the JSON reporter,
    # NOT result.status): "expected" means the run matched what the test
    # declared it should do - a plain pass, OR a test.fail()-annotated test
    # that correctly failed to confirm a documented defect. "unexpected"
    # means it did NOT match - either a real regression, or a
    # test.fail()-annotated defect that stopped reproducing (good news, but
    # needs the annotation removed) or timed out for the wrong reason.
    "expected": "Pass",
    "unexpected": "Fail",
    "flaky": "Flaky",
    "skipped": "Skipped",
    # Fallback keys in case a report ever exposes the raw result status
    # instead (defensive - should not normally be hit).
    "passed": "Pass",
    "failed": "Fail",
    "timedOut": "Fail",
    "interrupted": "Fail",
}


def walk_specs(suite, file_hint=None):
    """Yield (title, file, tests) for every spec under this suite, recursively."""
    file_hint = suite.get("file", file_hint)
    for spec in suite.get("specs", []):
        yield spec.get("title", ""), spec.get("file", file_hint), spec.get("tests", [])
    for sub in suite.get("suites", []):
        yield from walk_specs(sub, file_hint)


def last_result(test):
    results = test.get("results", [])
    if not results:
        return None
    return results[-1]


def load_run(results_path: Path):
    """Parse the Playwright JSON report into {tc_id: [entries]}."""
    data = json.loads(results_path.read_text(encoding="utf-8"))
    run_started = data.get("stats", {}).get("startTime", "unknown")

    entries_by_tc = {}
    unmatched_titles = []

    for suite in data.get("suites", []):
        for title, file, tests in walk_specs(suite):
            m = TC_ID_RE.match(title.strip())
            if not m:
                unmatched_titles.append(title)
                continue
            tc_id = m.group(1)
            sheet = FILE_TO_SHEET.get(Path(file).name if file else "", None)

            for test in tests:
                res = last_result(test)
                # test["status"] is the outcome category (expected/unexpected/
                # flaky/skipped) - this is what Status should be derived from,
                # since it correctly accounts for test.fail()/fixme()
                # annotations. res["status"] is the raw execution result
                # (passed/failed/timedOut/skipped) - used for the human
                # description in Actual Result, since "expected"/"unexpected"
                # alone doesn't say what actually happened.
                outcome_status = test.get("status") or "unknown"
                raw_status = res.get("status") if res else "unknown"
                duration = res.get("duration", 0) if res else 0
                error_msg = None
                if res and res.get("error"):
                    error_msg = strip_ansi(res["error"].get("message") or res["error"].get("stack"))
                fail_annotation = next(
                    (a.get("description") for a in test.get("annotations", []) if a.get("type") == "fail"),
                    None,
                )
                entries_by_tc.setdefault(tc_id, []).append(
                    {
                        "title": title,
                        "sheet": sheet,
                        "status": outcome_status,
                        "raw_status": raw_status,
                        "duration": duration,
                        "error": error_msg,
                        "fail_annotation": fail_annotation,
                    }
                )

    return entries_by_tc, unmatched_titles, run_started


def pick_canonical(entries):
    """Return (canonical_entry, superseded_entries). Prefers a '[DEF-' tagged title."""
    if len(entries) == 1:
        return entries[0], []
    def_entries = [e for e in entries if "[DEF-" in e["title"]]
    if def_entries:
        canonical = def_entries[-1]
    else:
        canonical = entries[-1]
    superseded = [e for e in entries if e is not canonical]
    return canonical, superseded


def format_actual_result(entry, run_label):
    raw_status = entry.get("raw_status", "unknown")
    outcome = entry.get("status", "unknown")
    duration = entry.get("duration", 0)
    is_annotated_fail = bool(entry.get("fail_annotation"))
    annotation = entry.get("fail_annotation")

    if raw_status == "passed" and not is_annotated_fail:
        return f"[Automated - {run_label}] PASSED ({duration} ms). Observed behavior matched Expected Result."

    if raw_status == "passed" and is_annotated_fail:
        # test.fail()-annotated (expects a rejection/defect to reproduce) but
        # actually passed -> the assertion for "still buggy" did NOT hold.
        # Either the defect got fixed, or the assertion needs revisiting.
        return (f"[Automated - {run_label}] UNEXPECTED PASS ({duration} ms) - test is annotated "
                f"test.fail() (\"{annotation}\") expecting the defect to still reproduce, but it "
                f"passed instead. Re-verify: defect may be fixed, or the assertion may not be "
                f"catching it correctly.")

    if raw_status in ("failed", "timedOut", "interrupted"):
        err = (entry.get("error") or "no error message captured").strip().splitlines()[0][:400]
        timeout_note = " (hit the global test timeout rather than a normal assertion failure - see Notes)" if raw_status == "timedOut" else ""
        if is_annotated_fail:
            return (f"[Automated - {run_label}] FAILED AS EXPECTED{timeout_note} ({duration} ms) - "
                    f"confirms defect (\"{annotation}\"). {err}")
        return f"[Automated - {run_label}] FAILED{timeout_note} ({duration} ms). {err}"

    if raw_status == "skipped":
        return f"[Automated - {run_label}] SKIPPED - not executed this run."

    return f"[Automated - {run_label}] outcome={outcome} raw_status={raw_status}"


def find_header(ws):
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value).strip() == "TC ID":
                headers = {}
                for cc in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=cc).value
                    if v:
                        headers[str(v).strip()] = cc
                return r, headers
    raise RuntimeError(f"Could not find 'TC ID' header row in sheet {ws.title}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", default="reports/test-results/latest/results.json")
    parser.add_argument("--xlsx", default="inputdata/Demoblaze_QA_TestCases.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Print what would change, write nothing.")
    args = parser.parse_args()

    results_path = Path(args.results)
    xlsx_path = Path(args.xlsx)

    if not results_path.exists():
        print(f"No results file at {results_path}. Run `npm run test:manual-suite` first.")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"No xlsx file at {xlsx_path}.")
        sys.exit(1)

    entries_by_tc, unmatched_titles, run_started = load_run(results_path)
    run_label = run_started if run_started != "unknown" else datetime.now().isoformat(timespec="seconds")

    # Resolve canonical vs superseded per TC ID
    canonical_by_tc = {}
    superseded_notes_by_tc = {}
    dup_warnings = []
    for tc_id, entries in entries_by_tc.items():
        canonical, superseded = pick_canonical(entries)
        canonical_by_tc[tc_id] = canonical
        if superseded:
            note = "; ".join(f'superseded automated test "{e["title"]}"' for e in superseded)
            superseded_notes_by_tc[tc_id] = f"[Automation note] {note} - using \"{canonical['title']}\" as authoritative ({run_label})."
            if not any("[DEF-" in e["title"] for e in entries):
                dup_warnings.append(tc_id)

    if not args.dry_run:
        backup_path = xlsx_path.with_name(
            f"{xlsx_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{xlsx_path.suffix}"
        )
        shutil.copy2(xlsx_path, backup_path)
        print(f"Backup written: {backup_path}")

    wb = openpyxl.load_workbook(xlsx_path)

    updated_counts = {"Login": 0, "Cart": 0, "API": 0}
    no_match_tc_ids = {"Login": [], "Cart": [], "API": []}

    for sheet_name in ("Login", "Cart", "API"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row, headers = find_header(ws)
        col_tcid = headers["TC ID"]
        col_actual = headers["Actual Result"]
        col_status = headers["Status"]
        col_notes = headers.get("Notes")

        for r in range(header_row + 1, ws.max_row + 1):
            tc_id = ws.cell(row=r, column=col_tcid).value
            if not tc_id:
                continue
            tc_id = str(tc_id).strip()
            canonical = canonical_by_tc.get(tc_id)
            if not canonical or canonical["sheet"] != sheet_name:
                no_match_tc_ids[sheet_name].append(tc_id)
                continue

            actual_text = format_actual_result(canonical, run_label)
            status_text = STATUS_MAP.get(canonical["status"], canonical["status"])

            # Defensive: strip any remaining control characters (not just ANSI
            # color codes) that openpyxl's IllegalCharacterError would reject,
            # so one malformed error message can't abort the whole sync run.
            actual_text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", actual_text)

            if args.dry_run:
                print(f"[{sheet_name}] {tc_id}: Status -> {status_text} | Actual -> {actual_text[:80]}...")
            else:
                ws.cell(row=r, column=col_actual, value=actual_text)
                ws.cell(row=r, column=col_status, value=status_text)
                if tc_id in superseded_notes_by_tc and col_notes:
                    existing = ws.cell(row=r, column=col_notes).value
                    note = superseded_notes_by_tc[tc_id]
                    ws.cell(row=r, column=col_notes, value=f"{existing} | {note}" if existing else note)
            updated_counts[sheet_name] += 1

    if not args.dry_run:
        wb.save(xlsx_path)
        print(f"Saved: {xlsx_path}")

    print("\n--- Summary ---")
    for sheet_name in ("Login", "Cart", "API"):
        print(f"{sheet_name}: updated {updated_counts[sheet_name]} row(s); "
              f"{len(no_match_tc_ids[sheet_name])} TC ID(s) with no result this run "
              f"(left as-is): {no_match_tc_ids[sheet_name]}")
    if dup_warnings:
        print(f"\nWARNING - duplicate TC IDs with NO [DEF-xxx] tag to disambiguate "
              f"(used the last one found, please review): {dup_warnings}")
    if unmatched_titles:
        print(f"\n{len(unmatched_titles)} test title(s) in the JSON report did not start with a recognizable "
              f"TC ID and were ignored. These are usually Regression & Performance internal tests "
              f"(no TC ID in manual sheet): {unmatched_titles[:5]}")

    print("\n✅ Done. Login/Cart/API results synced to xlsx. Regression/Performance results "
          "archived in reports/test-results/latest/results.json only (no TC ID mapping).")


if __name__ == "__main__":
    main()
